#!/anaconda3/envs/python_excel/bin/python3.6
# -*- coding: utf-8 -*-
"""
__title__ = '计算,汇总窗户尺寸'
__author__ = 'xiechaofan'
__mtime__ = '2018/7/29'
"""

from openpyxl import load_workbook

# 读取基础数据

rb = load_workbook("./窗料计算.xlsx")
sheet_basis = rb["基础数据"]
sheet_windows = rb["窗户数据"]

glasse_list = {}  # 玻璃尺寸汇总
baiye_list = {}  # 百叶尺寸汇总
pingkai_list = {}  # 平开窗尺寸汇总

# # 计算窗户数量
# b = -1
# for i in range(1, 500):
#     # a = rs.cell(row=i, column=4).value
#     if rs.cell(row=i, column=2).value:
#         b += 1

# 计算每扇窗户的起始行
c = []
for i in range(2, 500):
    # a = rs.cell(row=i, column=4).value
    if sheet_windows.cell(row=i, column=2).value:
        c.append(i)
print(c)
#
# def fact(n):
#     # 递归返回行号
#     if n == 1:
#         return 2
#     return fact(n - 1) + 6
#
#
# for i in range(1, b + 1):
#     c.append(fact(i))

# 单独计算每扇窗户的玻璃类型及数量,然后汇总
for a in c:
    width = []  # 宽度列表,从左至右
    height = []  # 高度列表,从上至下
    baiye_position = []  # 百叶位置列表,以两位数表示
    pingkai_position = []  # 平开窗位置列表,以两位数形式表示

    if sheet_windows.cell(row=a, column=3).value:
        area = sheet_windows.cell(row=a, column=2).value  # 区域及层数
        size = sheet_windows.cell(row=a, column=3).value  # 窗户编号,从左至右
        rows = sheet_windows.cell(row=a, column=4).value  # 总行数
        columns = sheet_windows.cell(row=a, column=5).value  # 总列数

        for i in range(a, a + 6):  # 获取当前编号窗户的尺寸区域，下一区域行号-当前窗户行号
            b = sheet_windows.cell(row=i, column=3).value
            if b:
                width.append(b)

        for i in range(a, a + 6):
            b = sheet_windows.cell(row=i, column=7).value
            if b:
                height.append(b)

        for i in range(a, a + 6):
            b = sheet_basis.cell(row=i, column=8).value
            if b:
                baiye_position.append(b)

        for i in range(a, a + 6):
            b = sheet_windows.cell(row=i, column=9).value
            if b:
                pingkai_position.append(b)

        # 计算单个尺寸
        for i in range(1, rows + 1):
            for j in range(1, columns + 1):
                position = i * 10 + j
                if position in pingkai_position:
                    glass_data = (width[j - 1], height[i - 1], "平开窗")
                    if glass_data in pingkai_list:
                        pingkai_list[glass_data] += 1
                    else:
                        pingkai_list[glass_data] = 1
                elif position in baiye_position:
                    glass_data = (width[j - 1] + 36, height[i - 1] + 36, "百叶")
                    if glass_data in baiye_list:
                        baiye_list[glass_data] += 1
                    else:
                        baiye_list[glass_data] = 1
                else:
                    glass_data = (width[j - 1] + 36, height[i - 1] + 36, "玻璃")
                    if glass_data in glasse_list:
                        glasse_list[glass_data] += 1
                    else:
                        glasse_list[glass_data] = 1

print(glasse_list)
print(baiye_list)
print(pingkai_list)

# 新建页面保存数据
# wb = Workbook()
# ws = wb.active

# 删除原有汇总页面
a = sheet_basis.cell(row=2, column=1).value
if a in rb.sheetnames:
    rb.remove(worksheet=rb[a])
else:
    pass
ws = rb.create_sheet(title=a)

# 将字典内容转换为列表数据
glasse_list2 = []
for i in glasse_list:
    glasse_list2.append([i[0], i[1], glasse_list[i]])

# 逐行写入数据
ws.cell(row=1, column=1).value = ws.title + "玻璃"
ws.cell(row=2, column=1).value = "宽"
ws.cell(row=2, column=2).value = "高"
ws.cell(row=2, column=3).value = "数量"

for j in range(len(glasse_list)):
    ws.cell(row=j + 3, column=1).value = glasse_list2[j][0]
    ws.cell(row=j + 3, column=2).value = glasse_list2[j][1]
    ws.cell(row=j + 3, column=3).value = glasse_list2[j][2]

baiye_list2 = []
for i in baiye_list:
    baiye_list2.append([i[0], i[1], baiye_list[i]])

ws.cell(row=1, column=6).value = ws.title + "百叶"
ws.cell(row=2, column=6).value = "宽"
ws.cell(row=2, column=7).value = "高"
ws.cell(row=2, column=8).value = "数量"

for j in range(len(baiye_list)):
    ws.cell(row=j + 3, column=6).value = baiye_list2[j][0]
    ws.cell(row=j + 3, column=7).value = baiye_list2[j][1]
    ws.cell(row=j + 3, column=8).value = baiye_list2[j][2]

pingkai_list2 = []
for i in pingkai_list:
    pingkai_list2.append([i[0], i[1], pingkai_list[i]])

ws.cell(row=1, column=11).value = ws.title + "平开窗洞口尺寸"
ws.cell(row=2, column=11).value = "宽"
ws.cell(row=2, column=12).value = "高"
ws.cell(row=2, column=13).value = "数量"

for j in range(len(pingkai_list)):
    ws.cell(row=j + 3, column=11).value = pingkai_list2[j][0]
    ws.cell(row=j + 3, column=12).value = pingkai_list2[j][1]
    ws.cell(row=j + 3, column=13).value = pingkai_list2[j][2]

# 保存表格
rb.save("./窗料计算.xlsx")
